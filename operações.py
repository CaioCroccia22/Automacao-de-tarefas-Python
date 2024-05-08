from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='files/gastos.xlsx')
planilha = wb['Sheet']
# print(planilha)

valor_total = 0
# 1 - Somando os valores
for i in range(2, 7):
    valor = int(planilha['B%s' %i].value)
    valor_total += valor
print(valor_total)

# Inserindo valor na planilha
planilha['B9'] = valor_total

# Salvando planilha
wb.save(filename='files/gastos.xlsx')

# Mesclar celulas
planilha['A8'] = 'Teste'
planilha.merge_cells('A8:B8')
wb.save(filename='files/gastos.xlsx')
    